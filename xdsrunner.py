import os
import openpyxl
import sys

def process_xds_folder(folder_path):
    # Traverse the selected folder and its subdirectories
    for root, dirs, files in os.walk(folder_path):
        # do something with each file
        pass
    
    # Search for xds.inp file (case-insensitive) in the selected folder and its subfolders
    xds_files = []
    for root, dirs, files in os.walk(folder_path):
        if "xds.inp" in [file.lower() for file in files]:
            xds_files.append(os.path.join(root, "xds.inp"))
    
    # Create and initialize excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "XDS Results"
    sheet["A1"] = "No."
    sheet["B1"] = "Path"
    sheet["C1"] = "Integation_cell"
    sheet["D1"] = "Space group"
    sheet["E1"] = "Unit cell"
    sheet["F1"] = "Isa"
    sheet["G1"] = "CC1/2"
    sheet["H1"] = "Completeness"
    sheet["I1"] = "Pseudo Resolution"
    
    # Initialize row number
    row_num = 2
    Number_counts = 0
    # Run xds in the same directory as each xds.inp file
    for i, xds_file in enumerate(xds_files):
        xds_dir = os.path.dirname(xds_file)
        os.chdir(xds_dir)
        os.system("xds_par")
        Number_counts = Number_counts + 1
        # Read results from XDS_ASCII.HKL file
        xds_ascii = os.path.join(xds_dir, "XDS_ASCII.HKL")
        if not os.path.isfile(xds_ascii):
            continue
        with open(xds_ascii) as f:
            lines = f.readlines()
            for j, line in enumerate(lines):
                if "!SPACE_GROUP_NUMBER=" in line:
                    space_group = line.split("=")[1].strip()
                    sheet.cell(row=row_num, column=4).value = space_group
                elif "!UNIT_CELL_CONSTANTS=" in line:
                    unit_cell = line.split("=")[1].strip()
                    sheet.cell(row=row_num, column=5).value = unit_cell
    
        # Read results from INTEGRATE.HKL file
        integrate_hkl = os.path.join(xds_dir, "INTEGRATE.HKL")
        with open(integrate_hkl) as f:
            lines = f.readlines()
            for line in lines:
                if "!UNIT_CELL_CONSTANTS=" in line:
                    p1_cell = line.split("=")[1].strip()
                    sheet.cell(row=row_num, column=3).value = p1_cell
                    
        # Read results from CORRECT.LP file
        correct_lp = os.path.join(xds_dir, "CORRECT.LP")
        resolution = None
        isa = None
        cc12 = None
        completeness = None
        with open(correct_lp) as f:
            lines = f.readlines()
            for j, line in enumerate(lines):
                if "WILSON STATISTICS OF DATA SET" in line:
                    cc12_str = lines[j-12].split()[10]
                    cc12 = float(cc12_str.replace('*', ''))
                    completeness_per = lines[j-12].split()[4]
                    completeness = float(completeness_per.replace('%', ''))
                    for i in range(1, len(lines)):
                        if j-i-12 < 0:
                            break
                        if "*" in lines[j-i-12]:
                            resolution_str = lines[j-i-12].split()[0]
                            resolution = resolution_str
                            break
                        elif j-i-13 < 0:
                            break
                        elif "*" in lines[j-i-13]:
                            resolution_str = lines[j-i-13].split()[0]
                            resolution = resolution_str
                            break
                                
        with open(correct_lp) as f:
            lines = f.readlines()            
            for j, line in enumerate(lines):        
                if "a        b          ISa" in line:
                    isa_str = lines[j+1].strip()
                    isa = float(isa_str.split()[2])
                    break
    
        # Write xds file name and path to excel sheet
        sheet.cell(row=row_num, column=1).value = Number_counts
        sheet.cell(row=row_num, column=2).value = xds_dir
        sheet.cell(row=row_num, column=7).value = cc12
        sheet.cell(row=row_num, column=8).value = completeness
        sheet.cell(row=row_num, column=9).value = resolution
        sheet.cell(row=row_num, column=6).value = isa
        
        # Increment row number
        row_num += 1
    
    
    # Auto adjust column width to fit content
    for col in sheet.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 1) * 1.2 # 1.2 is a scaling factor to adjust for font size
        sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = adjusted_width
    
    # Save excel workbook
    workbook.save(os.path.join(folder_path, "xdsrunner.xlsx"))
    
    
    print("All information extracted. Note that resolution is ideal.")
    
if __name__ == "__main__":
    if len(sys.argv) > 1:
        folder_path = sys.argv[1]
        print(f"xdsrunner has received input path: {folder_path}")
        process_xds_folder(folder_path)
    else:
        print("No input path provided. Exiting...")
        exit()